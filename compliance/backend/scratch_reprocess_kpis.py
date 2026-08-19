import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.append(os.path.dirname(os.path.abspath(__file__)))
from core.database import SessionLocal
from models.all_models import KPIEvaluation, KPIEvaluationDetail, Evidence

def safe_float(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)): return float(val)
    s = str(val).strip().replace(',', '').replace('$', '')
    is_pct = False
    if s.endswith('%'):
        s = s[:-1].strip()
        is_pct = True
    try:
        f = float(s)
        return f / 100.0 if is_pct else f
    except (ValueError, TypeError): 
        return 0.0

def process_kpis():
    db = SessionLocal()
    evaluations = db.query(KPIEvaluation).all()
    
    for evaluation in evaluations:
        evidence = db.query(Evidence).filter(Evidence.kpi_eval_id == evaluation.id).first()
        if not evidence or not evidence.file_path.endswith((".xlsx", ".xlsm")):
            print(f"  Skipping {evaluation.department} (no xlsx)")
            continue
            
        file_path = evidence.file_path
        if not os.path.exists(file_path):
            print(f"  File not found: {file_path}")
            continue
            
        print(f"\nProcessing: {evaluation.department} | Period: {evaluation.period_month}/{evaluation.period_year}")
        
        # Clear existing details
        db.query(KPIEvaluationDetail).filter(KPIEvaluationDetail.evaluation_id == evaluation.id).delete()
        
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = None
            for sheet in wb.sheetnames:
                if "Tablero General" in sheet:
                    ws = wb[sheet]
                    break
            if not ws:
                ws = wb.worksheets[0]
            
            print(f"  Using sheet: {ws.title}")
            
            header_map = {}
            total_score = 0.0
            
            for row in ws.iter_rows(values_only=True):
                if not row: continue
                
                # Buscamos la fila de encabezados si no la tenemos
                if not header_map:
                    candidate = {}
                    for i, cell in enumerate(row):
                        val = str(cell).lower().strip() if cell else ""
                        if val == "#" or val == "no." or val == "num": candidate["num"] = i
                        elif "kpi" == val or "indicador" in val or "kpis" == val: candidate["kpi"] = i
                        elif "frecuencia" in val: candidate["freq"] = i
                        elif "meta" in val: candidate["target"] = i
                        elif "ponderación" in val or "ponderacion" in val or "ideal" == val: candidate["weight"] = i
                        elif "cumplimiento ponderado" in val or "evaluacion" == val: candidate["comp_weight"] = i
                        elif "cumplimiento" in val: 
                            if "comp_month" not in candidate: candidate["comp_month"] = i
                        elif "estado" in val: candidate["status"] = i
                    # Aceptar si encontramos al menos 3 columnas conocidas
                    if len(candidate) >= 3:
                        header_map = candidate
                        print(f"  Headers found: {header_map}")
                    continue
                
                # Si tenemos headers, buscar filas de KPIs
                num_idx = header_map.get("num")
                kpi_idx = header_map.get("kpi")
                
                is_kpi_row = False
                if num_idx is not None and num_idx < len(row) and str(row[num_idx]).isdigit():
                    is_kpi_row = True
                elif kpi_idx is not None and kpi_idx < len(row) and isinstance(row[kpi_idx], str) and row[kpi_idx].strip() and row[0] is None:
                    pass
                
                if is_kpi_row:
                    def get_val(key, default=""):
                        idx = header_map.get(key)
                        return row[idx] if idx is not None and idx < len(row) and row[idx] is not None else default
                    
                    kpi_name = str(get_val("kpi"))
                    freq = str(get_val("freq"))
                    target = str(get_val("target"))
                    weight = safe_float(get_val("weight"))
                    comp_month = safe_float(get_val("comp_month"))
                    comp_weight = safe_float(get_val("comp_weight"))
                    status_kpi = str(get_val("status")).strip()
                    
                    # Calcular comp_month o comp_weight si faltan
                    # Fix user data entry error: if they put the weighted score (e.g. 0.09) in the monthly score column for a weight of 0.10
                    if comp_month > 0 and comp_month <= weight and weight < 1.0 and comp_weight == 0:
                        comp_weight = comp_month
                        comp_month = comp_weight / weight
                    elif comp_weight > 0 and comp_month == 0 and weight > 0:
                        comp_month = comp_weight / weight
                    elif comp_month > 0 and comp_weight == 0 and weight > 0:
                        comp_weight = comp_month * weight

                    # Autocalcular estado forzando la lógica matemática para evitar errores humanos en el Excel
                    if weight > 0:
                        cm_rounded = round(comp_month, 4)
                        if cm_rounded >= 0.9: status_kpi = "✅ Cumple"
                        elif cm_rounded >= 0.7: status_kpi = "⚠️ En riesgo"
                        else: status_kpi = "❌ Incumple"
                    
                    if kpi_name:
                        total_score += comp_weight
                        detail = KPIEvaluationDetail(
                            evaluation_id=evaluation.id,
                            kpi_name=kpi_name,
                            frequency=freq,
                            target=target,
                            weight=weight,
                            compliance_month=comp_month,
                            compliance_weighted=comp_weight,
                            status=status_kpi
                        )
                        db.add(detail)
                        print(f"    KPI: {kpi_name[:40]:40s} | W: {weight:.2f} | C: {comp_month:.2f} | S: {status_kpi}")
                
                # Revisar si hay una fila explicita de calificacion global
                found_global = False
                for cell in row[:3]:
                    if cell and "CALIFICACI" in str(cell).upper() and "GLOBAL" in str(cell).upper():
                        found_global = True
                        break
                if found_global:
                    comp_idx = header_map.get("comp_month") or header_map.get("comp_weight")
                    if comp_idx and comp_idx < len(row) and row[comp_idx] is not None:
                        evaluation.global_score = safe_float(row[comp_idx])
                        total_score = evaluation.global_score
                        print(f"  >>> Global Score: {evaluation.global_score:.4f} ({evaluation.global_score*100:.1f}%)")
            
            # Si no hubo fila explicita de global score (ej. RH), usamos la suma de pesos ponderados
            if not found_global and total_score > 0:
                evaluation.global_score = total_score
                print(f"  >>> Computed Global Score: {evaluation.global_score:.4f}")
            
            print(f"  Updated global_score = {evaluation.global_score}")
                    
        except Exception as e:
            print(f"  ERROR: {e}")
            
    db.commit()
    db.close()
    print("\n=== DONE ===")

if __name__ == "__main__":
    process_kpis()
