import sys
import os
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
sys.path.append(os.path.dirname(os.path.abspath(__file__)))

# Check contabilidad file structure
file_path = r'uploads\kpis\3236a08822a84da1bcf794a765f5fb23.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
print("=== CONTABILIDAD ===")
print("Sheets:", wb.sheetnames)
ws = None
for sheet in wb.sheetnames:
    if "Tablero General" in sheet:
        ws = wb[sheet]
        break
if not ws:
    ws = wb.worksheets[0]
print(f"Using sheet: {ws.title}")
for i, row in enumerate(ws.iter_rows(values_only=True)):
    if any(cell is not None for cell in row):
        print(f"Row {i}: {row}")
    if i > 25:
        break
