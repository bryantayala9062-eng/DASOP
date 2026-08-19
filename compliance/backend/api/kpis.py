from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File, Form
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from datetime import datetime, timezone
import os
import uuid
import openpyxl
import logging

logger = logging.getLogger("complianceop")

from core.database import get_db
from models.all_models import User, KPIEvaluation, Evidence, KPIEvaluationDetail
from api.auth import get_current_user

router = APIRouter(prefix="/api/kpis", tags=["kpis"])

UPLOAD_DIR = "uploads/kpis"
os.makedirs(UPLOAD_DIR, exist_ok=True)

@router.post("/evaluation/upload")
async def upload_kpi_evaluation(
    department: str = Form(...),
    period_month: int = Form(...),
    period_year: int = Form(...),
    global_score: float = Form(...),
    collaborator_name: Optional[str] = Form(None),
    comments: Optional[str] = Form(None),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    if current_user.role != "admin" and department != current_user.department:
        raise HTTPException(status_code=403, detail="No tienes permiso para evaluar KPIs de otro departamento")

    # Read and save file
    file_ext = os.path.splitext(file.filename)[1]
    safe_filename = f"{uuid.uuid4().hex}{file_ext}"
    file_path = os.path.join(UPLOAD_DIR, safe_filename)

    with open(file_path, "wb") as f:
        content = await file.read()
        f.write(content)

    file_size = len(content)

    # Check existing evaluation
    existing_eval = db.query(KPIEvaluation).filter(
        KPIEvaluation.department == department,
        KPIEvaluation.period_month == period_month,
        KPIEvaluation.period_year == period_year
    ).first()

    if existing_eval:
        existing_eval.global_score = global_score
        existing_eval.collaborator_name = collaborator_name
        existing_eval.comments = comments
        existing_eval.evaluated_by_user_id = current_user.id
        existing_eval.evaluation_date = datetime.now(timezone.utc)
        evaluation = existing_eval
        # Clean old evidence (optional, we could keep it, but lets replace)
        old_evidence = db.query(Evidence).filter(Evidence.kpi_eval_id == evaluation.id).all()
        for ev in old_evidence:
            try:
                os.remove(ev.file_path)
            except Exception as e:
                logger.warning(f"No se pudo eliminar el archivo de evidencia anterior ({ev.file_path}): {e}")
        db.query(Evidence).filter(Evidence.kpi_eval_id == evaluation.id).delete()
        db.query(KPIEvaluationDetail).filter(KPIEvaluationDetail.evaluation_id == evaluation.id).delete()
    else:
        evaluation = KPIEvaluation(
            department=department,
            period_month=period_month,
            period_year=period_year,
            global_score=global_score,
            collaborator_name=collaborator_name,
            comments=comments,
            evaluated_by_user_id=current_user.id
        )
        db.add(evaluation)
        db.flush()

    # Parse Excel for KPI Details
    if file_ext in [".xlsx", ".xlsm"]:
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            ws = None
            for sheet in wb.sheetnames:
                if "Tablero General" in sheet:
                    ws = wb[sheet]
                    break
            if not ws:
                ws = wb.worksheets[0]
            
            header_map = {}
            total_score = 0.0
            for row in ws.iter_rows(values_only=True):
                if not row: continue
                
                # Buscamos la fila de encabezados si no la tenemos
                if not header_map:
                    candidate = {}
                    for i, cell in enumerate(row):
                        val = str(cell).lower().strip() if cell else ""
                        if val == "#" or val == "no." or val == "num": candidate["num"] = i
                        elif "kpi" == val or "indicador" in val or "kpis" == val: candidate["kpi"] = i
                        elif "frecuencia" in val: candidate["freq"] = i
                        elif "meta" in val: candidate["target"] = i
                        elif "ponderación" in val or "ponderacion" in val or "ideal" == val: candidate["weight"] = i
                        elif "cumplimiento ponderado" in val or "evaluacion" == val: candidate["comp_weight"] = i
                        elif "cumplimiento" in val: 
                            if "comp_month" not in candidate: candidate["comp_month"] = i
                        elif "estado" in val: candidate["status"] = i
                    # Aceptar si encontramos al menos 3 columnas conocidas
                    if len(candidate) >= 3:
                        header_map = candidate
                    continue
                
                # Si tenemos headers, buscar filas de KPIs (identificadas porque la columna '#' tiene un número, o la fila tiene un KPI válido)
                num_idx = header_map.get("num")
                kpi_idx = header_map.get("kpi")
                
                is_kpi_row = False
                if num_idx is not None and num_idx < len(row) and str(row[num_idx]).isdigit():
                    is_kpi_row = True
                elif kpi_idx is not None and kpi_idx < len(row) and isinstance(row[kpi_idx], str) and row[kpi_idx].strip() and row[0] is None:
                    # Alternativa: no hay numero pero hay nombre de KPI (a veces pasa en formatos raros)
                    pass
                
                if is_kpi_row:
                    def get_val(key, default=""):
                        idx = header_map.get(key)
                        return row[idx] if idx is not None and idx < len(row) and row[idx] is not None else default
                    
                    kpi_name = str(get_val("kpi"))
                    freq = str(get_val("freq"))
                    target = str(get_val("target"))
                    
                    def safe_float(val):
                        if val is None: return 0.0
                        if isinstance(val, (int, float)): return float(val)
                        s = str(val).strip().replace(',', '').replace('$', '')
                        is_pct = False
                        if s.endswith('%'):
                            s = s[:-1].strip()
                            is_pct = True
                        try:
                            f = float(s)
                            return f / 100.0 if is_pct else f
                        except (ValueError, TypeError): 
                            return 0.0
                    
                    weight = safe_float(get_val("weight"))
                    comp_month = safe_float(get_val("comp_month"))
                    comp_weight = safe_float(get_val("comp_weight"))
                    status_kpi = str(get_val("status")).strip()
                    
                    # Calcular comp_month o comp_weight si faltan
                    # Fix user data entry error: if they put the weighted score (e.g. 0.09) in the monthly score column for a weight of 0.10
                    if comp_month > 0 and comp_month <= weight and weight < 1.0 and comp_weight == 0:
                        comp_weight = comp_month
                        comp_month = comp_weight / weight
                    elif comp_weight > 0 and comp_month == 0 and weight > 0:
                        comp_month = comp_weight / weight
                    elif comp_month > 0 and comp_weight == 0 and weight > 0:
                        comp_weight = comp_month * weight

                    # Autocalcular estado forzando la lógica matemática para evitar errores humanos en el Excel
                    if weight > 0:
                        cm_rounded = round(comp_month, 4)
                        if cm_rounded >= 0.9: status_kpi = "✅ Cumple"
                        elif cm_rounded >= 0.7: status_kpi = "⚠️ En riesgo"
                        else: status_kpi = "❌ Incumple"
                    
                    if kpi_name:
                        total_score += comp_weight
                        detail = KPIEvaluationDetail(
                            evaluation_id=evaluation.id,
                            kpi_name=kpi_name,
                            frequency=freq,
                            target=target,
                            weight=weight,
                            compliance_month=comp_month,
                            compliance_weighted=comp_weight,
                            status=status_kpi
                        )
                        db.add(detail)
                
                # Revisar si hay una fila explicita de calificacion global
                found_global = False
                for cell in row[:3]:
                    if cell and "CALIFICACI" in str(cell).upper() and "GLOBAL" in str(cell).upper():
                        found_global = True
                        break
                if found_global:
                    comp_idx = header_map.get("comp_month") or header_map.get("comp_weight")
                    if comp_idx and comp_idx < len(row) and row[comp_idx] is not None:
                        evaluation.global_score = safe_float(row[comp_idx])
                        total_score = evaluation.global_score # override computed total
            
            # Si no hubo fila explicita de global score (ej. RH), usamos la suma de pesos ponderados
            if not found_global and total_score > 0:
                evaluation.global_score = total_score

            db.flush()
        except Exception as e:
            logger.error(f"Error procesando Excel de KPIs: {e}")

    # Create evidence
    evidence = Evidence(
        title=f"Reporte KPI {department} {period_month}/{period_year}",
        file_path=file_path,
        file_name=file.filename,
        file_type=file.content_type,
        file_size=file_size,
        kpi_eval_id=evaluation.id,
        category="KPI_Report",
        uploaded_by_user_id=current_user.id,
        status="reviewed"
    )
    db.add(evidence)
    db.commit()

    return {"message": "Evaluación de KPIs subida correctamente", "id": evaluation.id}

@router.get("/evaluation/history")
def get_kpi_history(
    department: Optional[str] = None,
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    query = db.query(KPIEvaluation).options(joinedload(KPIEvaluation.evidences), joinedload(KPIEvaluation.evaluator), joinedload(KPIEvaluation.details))
    
    if current_user.role != "admin":
        department = current_user.department
        
    if department and department != "undefined" and department != "null" and department != "":
        query = query.filter(KPIEvaluation.department == department)
        
    if year:
        query = query.filter(KPIEvaluation.period_year == year)
        
    evaluations = query.order_by(KPIEvaluation.period_month.desc()).all()
    
    result = []
    for e in evaluations:
        result.append({
            "id": e.id,
            "department": e.department,
            "period_month": e.period_month,
            "period_year": e.period_year,
            "global_score": e.global_score,
            "collaborator_name": e.collaborator_name,
            "comments": e.comments,
            "evaluation_date": e.evaluation_date.isoformat() if e.evaluation_date else None,
            "evaluator_name": e.evaluator.full_name if e.evaluator else "Desconocido",
            "details": [{
                "kpi_name": d.kpi_name,
                "weight": d.weight,
                "compliance_month": d.compliance_month,
                "status": d.status
            } for d in e.details],
            "evidence": {
                "id": e.evidences[0].id,
                "file_name": e.evidences[0].file_name,
                "file_path": f"/api/evidences/{e.evidences[0].id}/download"
            } if e.evidences else None
        })
        
    return result


@router.delete("/evaluation/{eval_id}", status_code=204)
def delete_kpi_evaluation(
    eval_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Elimina una evaluación de KPIs y sus evidencias. Solo admin."""
    evaluation = db.query(KPIEvaluation).filter(KPIEvaluation.id == eval_id).first()
    if not evaluation:
        raise HTTPException(status_code=404, detail="Evaluación no encontrada")

    if current_user.role != "admin" and evaluation.department != current_user.department:
        raise HTTPException(status_code=403, detail="Solo administradores pueden eliminar evaluaciones de otros departamentos")

    # Eliminar evidencias asociadas
    evidences = db.query(Evidence).filter(Evidence.kpi_eval_id == eval_id).all()
    for ev in evidences:
        if ev.file_path and os.path.exists(ev.file_path):
            try:
                os.remove(ev.file_path)
            except Exception:
                pass
        db.delete(ev)

    db.delete(evaluation)
    db.commit()

    return None


@router.get("/executive-summary")
def get_executive_summary(
    year: Optional[int] = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Dashboard ejecutivo: scores por área, tendencias y KPIs críticos."""
    from datetime import datetime, timezone
    if not year:
        year = datetime.now(timezone.utc).year

    evaluations = db.query(KPIEvaluation).options(
        joinedload(KPIEvaluation.details)
    ).filter(
        KPIEvaluation.period_year == year
    ).order_by(KPIEvaluation.department, KPIEvaluation.period_month).all()

    dept_data = {}
    critical_kpis = []

    for e in evaluations:
        if e.department not in dept_data:
            dept_data[e.department] = []

        details_list = []
        for d in e.details:
            detail_obj = {
                "kpi_name": d.kpi_name,
                "weight": d.weight or 0,
                "compliance_month": d.compliance_month or 0,
                "compliance_weighted": d.compliance_weighted or 0,
                "status": d.status or ""
            }
            details_list.append(detail_obj)

            # Identify critical KPIs (not "cumple" or "Cumple", or explicitly "incumple")
            status_lower = (d.status or "").lower().strip()
            is_good = "cumple" in status_lower and "incumple" not in status_lower
            if status_lower and not is_good:
                critical_kpis.append({
                    "department": e.department,
                    "kpi_name": d.kpi_name,
                    "compliance": round((d.compliance_month or 0) * 100, 1),
                    "weight": round((d.weight or 0) * 100, 1),
                    "status": d.status,
                    "month": e.period_month
                })

        dept_data[e.department].append({
            "month": e.period_month,
            "global_score": round((e.global_score or 0) * 100, 1),
            "details": details_list
        })

    departments = []
    for dept_name, scores in dept_data.items():
        departments.append({
            "name": dept_name,
            "scores": scores
        })

    # Sort critical KPIs by compliance ascending (worst first)
    critical_kpis.sort(key=lambda x: x["compliance"])

    return {
        "year": year,
        "departments": departments,
        "critical_kpis": critical_kpis[:20]  # Top 20 worst
    }
