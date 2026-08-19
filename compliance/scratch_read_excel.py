import sys
import json
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')

file_path = r'c:\Users\Administrador\Desktop\DashOP\compliance\backend\uploads\kpis\3236a08822a84da1bcf794a765f5fb23.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
sheets = wb.sheetnames
print("Sheets:", sheets)

for sheet_name in sheets:
    print(f"\n--- Sheet: {sheet_name} ---")
    ws = wb[sheet_name]
    # Read first 20 rows and 10 columns
    for row in ws.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10, values_only=True):
        if any(cell is not None for cell in row):
            print(row)
