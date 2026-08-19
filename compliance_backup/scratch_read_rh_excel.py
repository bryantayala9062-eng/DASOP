import sys
import openpyxl

sys.stdout.reconfigure(encoding='utf-8')
file_path = r'c:\Users\Administrador\Desktop\DashOP\compliance\backend\uploads\kpis\d30b436633264544a7c7a4039ea20eeb.xlsx'
wb = openpyxl.load_workbook(file_path, data_only=True)
print("Sheets:", wb.sheetnames)
ws = wb.worksheets[0]
print("--- First Sheet ---")
for row in ws.iter_rows(values_only=True):
    if any(cell is not None for cell in row):
        print(row)
